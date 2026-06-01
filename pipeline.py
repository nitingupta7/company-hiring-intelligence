import csv
import json
import logging
import os
import time
import zipfile
from pathlib import Path

try:
    from dotenv import load_dotenv
except ImportError:  # Keep the pipeline usable even before dependencies are installed.
    load_dotenv = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None


ROOT_DIR = Path(__file__).resolve().parent
INPUT_CSV = ROOT_DIR / "input" / "investments_VC.csv"
INPUT_ZIP = ROOT_DIR / "startup-investments-crunchbase.zip"
SOCIAL_CSV = ROOT_DIR / "output" / "step3_social.csv"
OUTPUT_DIR = ROOT_DIR / "output"
LOG_DIR = ROOT_DIR / "logs"
OUTPUT_XLSX = OUTPUT_DIR / "companies.xlsx"
OUTPUT_JSON = OUTPUT_DIR / "companies.json"
LOG_FILE = LOG_DIR / "pipeline.log"
API_DELAY_SECONDS = 2

OUTPUT_COLUMNS = [
    "Company Name",
    "Industry",
    "Total Funding",
    "Funding Stage",
    "Headquarters",
    "Website",
    "Careers Page",
    "Hiring Status",
    "LinkedIn URL",
    "X/Twitter URL",
    "Facebook URL",
    "Instagram URL",
    "YouTube URL",
    "Source URL",
]

ALIASES = {
    "Company Name": ["Company Name", "company_name", "name", "Name", "Organization Name"],
    "Industry": ["Industry", "industry", "market", " market ", "category", "category_list"],
    "Total Funding": [
        "Total Funding",
        "total_funding",
        "funding_total_usd",
        " funding_total_usd ",
        "Funding",
    ],
    "Funding Stage": ["Funding Stage", "funding_stage", "stage", "YC Batch"],
    "Headquarters": ["Headquarters", "headquarters", "location", "Location"],
    "Website": ["Website", "website", "homepage_url", "homepage", "url"],
    "Careers Page": ["Careers Page", "careers_page", "careers_url", "Careers URL"],
    "Hiring Status": ["Hiring Status", "hiring_status", "is_hiring", "Hiring"],
    "LinkedIn URL": ["LinkedIn URL", "LinkedIn", "linkedin", "linkedin_url"],
    "X/Twitter URL": [
        "X/Twitter URL",
        "Twitter URL",
        "twitter",
        "twitter_url",
        "x_url",
        "X URL",
    ],
    "Facebook URL": ["Facebook URL", "Facebook", "facebook", "facebook_url"],
    "Instagram URL": ["Instagram URL", "Instagram", "instagram", "instagram_url"],
    "YouTube URL": ["YouTube URL", "Youtube URL", "YouTube", "youtube", "youtube_url"],
    "Source URL": ["Source URL", "source_url", "yc_url", "permalink", "source"],
}


def setup_logging() -> None:
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s",
        handlers=[
            logging.FileHandler(LOG_FILE),
            logging.StreamHandler(),
        ],
    )


def load_environment() -> None:
    if load_dotenv is None:
        logging.warning("python-dotenv is not installed; skipping .env loading.")
        return

    env_path = ROOT_DIR / ".env"
    load_dotenv(env_path)
    loaded_keys = [
        key
        for key in ("TWITTER_BEARER_TOKEN", "YOUTUBE_API_KEY", "FB_APP_TOKEN", "WELLFOUND_API_KEY")
        if os.getenv(key)
    ]
    logging.info("Loaded .env file. API keys available: %s", ", ".join(loaded_keys) or "none")


def simulate_api_delay() -> None:
    """Central delay hook for future API calls; keeps the pipeline compliant today."""
    time.sleep(API_DELAY_SECONDS)


def ensure_input_csv() -> Path:
    INPUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    if INPUT_CSV.exists() and INPUT_CSV.stat().st_size > 0:
        return INPUT_CSV

    if INPUT_ZIP.exists():
        logging.info("input/investments_VC.csv not found. Extracting it from %s", INPUT_ZIP.name)
        with zipfile.ZipFile(INPUT_ZIP) as archive:
            archive.extract("investments_VC.csv", INPUT_CSV.parent)
        return INPUT_CSV

    raise FileNotFoundError(
        "Could not find input/investments_VC.csv or startup-investments-crunchbase.zip"
    )


def normalize_header(value: str) -> str:
    return (value or "").strip().lower().replace("_", " ")


def value_from(row: dict, aliases: list[str]) -> str:
    normalized = {normalize_header(key): value for key, value in row.items()}
    for alias in aliases:
        value = normalized.get(normalize_header(alias))
        if has_value(value):
            return clean_value(value)
    return "N/A"


def has_value(value) -> bool:
    if value is None:
        return False
    text = str(value).strip()
    return bool(text) and text.lower() not in {"nan", "none", "null", "n/a", "na", "-"}


def clean_value(value) -> str:
    if not has_value(value):
        return "N/A"
    text = str(value).strip()
    text = " ".join(text.split())
    return text if text else "N/A"


def clean_industry(row: dict) -> str:
    market = value_from(row, ["market", " market ", "Industry", "industry"])
    if market != "N/A":
        return market.strip(" |")

    category_list = value_from(row, ["category_list"])
    if category_list == "N/A":
        return "N/A"
    categories = [item.strip() for item in category_list.split("|") if item.strip()]
    return categories[0] if categories else "N/A"


def clean_funding(value: str) -> str:
    value = clean_value(value)
    if value == "N/A":
        return "N/A"
    digits = "".join(ch for ch in value if ch.isdigit() or ch == ".")
    if not digits:
        return "N/A"
    try:
        amount = float(digits)
    except ValueError:
        return value
    if amount.is_integer():
        return str(int(amount))
    return str(amount)


def build_headquarters(row: dict) -> str:
    direct = value_from(row, ["Headquarters", "headquarters", "location"])
    if direct != "N/A":
        return direct

    parts = [
        value_from(row, ["city"]),
        value_from(row, ["state_code"]),
        value_from(row, ["country_code"]),
    ]
    parts = [part for part in parts if part != "N/A"]
    return ", ".join(parts) if parts else "N/A"


def infer_funding_stage(row: dict) -> str:
    direct = value_from(row, ALIASES["Funding Stage"])
    if direct != "N/A":
        return direct

    ordered_rounds = [
        ("round_H", "Series H"),
        ("round_G", "Series G"),
        ("round_F", "Series F"),
        ("round_E", "Series E"),
        ("round_D", "Series D"),
        ("round_C", "Series C"),
        ("round_B", "Series B"),
        ("round_A", "Series A"),
        ("private_equity", "Private Equity"),
        ("angel", "Angel"),
        ("seed", "Seed"),
        ("venture", "Venture"),
    ]
    normalized = {normalize_header(key): value for key, value in row.items()}
    for key, label in ordered_rounds:
        raw_value = normalized.get(normalize_header(key), "0")
        try:
            if float(str(raw_value).replace(",", "").strip() or 0) > 0:
                return label
        except ValueError:
            continue
    return "N/A"


def normalize_hiring_status(value: str) -> str:
    value = clean_value(value)
    if value == "N/A":
        return "Unknown"

    lowered = value.lower()
    if lowered in {"hiring", "yes", "true", "1", "active", "now hiring"}:
        return "Hiring"
    if lowered in {"not hiring", "no", "false", "0", "inactive"}:
        return "Not Hiring"
    if "not hiring" in lowered:
        return "Not Hiring"
    if "hiring" in lowered or "open" in lowered:
        return "Hiring"
    return "Unknown"


def normalize_company_key(name: str) -> str:
    return "".join(ch.lower() for ch in clean_value(name) if ch.isalnum())


def source_url_from(row: dict) -> str:
    source = value_from(row, ALIASES["Source URL"])
    if source == "N/A":
        return "N/A"
    if source.startswith("/organization/"):
        return f"https://www.crunchbase.com{source}"
    return source


def read_csv_rows(path: Path) -> list[dict]:
    if not path.exists() or path.stat().st_size == 0:
        logging.warning("Skipping missing or empty CSV: %s", path)
        return []

    encodings = ("utf-8-sig", "utf-8", "latin-1")
    for encoding in encodings:
        try:
            with path.open("r", encoding=encoding, newline="") as file:
                rows = list(csv.DictReader(file))
            logging.info("Loaded %s rows from %s", len(rows), path)
            return rows
        except UnicodeDecodeError:
            continue
        except Exception:
            logging.exception("Failed reading CSV: %s", path)
            return []

    logging.error("Could not decode CSV: %s", path)
    return []


def base_record(row: dict) -> dict:
    return {
        "Company Name": value_from(row, ALIASES["Company Name"]),
        "Industry": clean_industry(row),
        "Total Funding": clean_funding(value_from(row, ALIASES["Total Funding"])),
        "Funding Stage": infer_funding_stage(row),
        "Headquarters": build_headquarters(row),
        "Website": value_from(row, ALIASES["Website"]),
        "Careers Page": "N/A",
        "Hiring Status": "Unknown",
        "LinkedIn URL": "N/A",
        "X/Twitter URL": "N/A",
        "Facebook URL": "N/A",
        "Instagram URL": "N/A",
        "YouTube URL": "N/A",
        "Source URL": source_url_from(row),
    }


def social_record(row: dict) -> dict:
    return {
        "Company Name": value_from(row, ALIASES["Company Name"]),
        "Industry": clean_industry(row),
        "Total Funding": clean_funding(value_from(row, ALIASES["Total Funding"])),
        "Funding Stage": infer_funding_stage(row),
        "Headquarters": build_headquarters(row),
        "Website": value_from(row, ALIASES["Website"]),
        "Careers Page": value_from(row, ALIASES["Careers Page"]),
        "Hiring Status": normalize_hiring_status(value_from(row, ALIASES["Hiring Status"])),
        "LinkedIn URL": value_from(row, ALIASES["LinkedIn URL"]),
        "X/Twitter URL": value_from(row, ALIASES["X/Twitter URL"]),
        "Facebook URL": value_from(row, ALIASES["Facebook URL"]),
        "Instagram URL": value_from(row, ALIASES["Instagram URL"]),
        "YouTube URL": value_from(row, ALIASES["YouTube URL"]),
        "Source URL": source_url_from(row),
    }


def merge_records(base_rows: list[dict], social_rows: list[dict]) -> list[dict]:
    merged: dict[str, dict] = {}

    for row in base_rows:
        record = base_record(row)
        key = normalize_company_key(record["Company Name"])
        if not key:
            continue
        merged[key] = record

    logging.info("Base records after normalization: %s", len(merged))

    for row in social_rows:
        record = social_record(row)
        key = normalize_company_key(record["Company Name"])
        if not key:
            continue
        existing = merged.get(key, {column: "N/A" for column in OUTPUT_COLUMNS})
        for column in OUTPUT_COLUMNS:
            if column == "Hiring Status":
                if record[column] != "Unknown":
                    existing[column] = record[column]
                elif existing.get(column, "Unknown") not in {"Hiring", "Not Hiring"}:
                    existing[column] = "Unknown"
                continue
            if has_value(record[column]):
                existing[column] = record[column]
        merged[key] = existing

    final_records = []
    for record in merged.values():
        final = {}
        for column in OUTPUT_COLUMNS:
            value = record.get(column, "N/A")
            if column == "Hiring Status":
                final[column] = normalize_hiring_status(value)
            else:
                final[column] = clean_value(value)
        final_records.append(final)

    final_records.sort(key=lambda item: item["Company Name"].lower())
    logging.info("Final merged records: %s", len(final_records))
    return final_records


def write_json(records: list[dict]) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    existing = json.load(OUTPUT_JSON.open()) if OUTPUT_JSON.exists() else []
    existing_names = {r.get("Company Name") for r in existing}
    new_records = [r for r in records if r.get("Company Name") not in existing_names]
    all_records = existing + new_records
    with OUTPUT_JSON.open("w", encoding="utf-8") as file:
        json.dump(all_records, file, indent=2, ensure_ascii=False)
    logging.info("Saved JSON output: %s", OUTPUT_JSON)


def write_xlsx(records: list[dict]) -> None:
    if OUTPUT_JSON.exists():
        import json as _json
        records = _json.load(OUTPUT_JSON.open())
    if Workbook is None:
        raise ImportError("openpyxl is required to write output/companies.xlsx")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Companies"
    worksheet.append(OUTPUT_COLUMNS)

    for record in records:
        worksheet.append([record[column] for column in OUTPUT_COLUMNS])

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for column_index, column_name in enumerate(OUTPUT_COLUMNS, start=1):
        max_length = len(column_name)
        for cell in worksheet.iter_rows(
            min_row=2,
            min_col=column_index,
            max_col=column_index,
            values_only=True,
        ):
            max_length = max(max_length, len(str(cell[0] or "")))
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max_length + 2, 60)

    worksheet.freeze_panes = "A2"
    workbook.save(OUTPUT_XLSX)
    logging.info("Saved Excel output: %s", OUTPUT_XLSX)


def main() -> int:
    setup_logging()
    logging.info("Starting company data pipeline")
    load_environment()

    try:
        input_csv = ensure_input_csv()
        simulate_api_delay()

        base_rows = read_csv_rows(input_csv)[:20]
        simulate_api_delay()

        social_rows = read_csv_rows(SOCIAL_CSV)
        simulate_api_delay()

        records = merge_records(base_rows, social_rows)
        if not records:
            logging.error("No records were produced. Check input files.")
            return 1

        write_json(records)
        write_xlsx(records)
        logging.info("Pipeline finished successfully")
        return 0
    except Exception:
        logging.exception("Pipeline failed")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
